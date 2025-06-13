import os
import django

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.sys.path.append(BASE_DIR)
os.environ['DJANGO_SETTINGS_MODULE'] = 'core.settings'
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")
django.setup()

import openpyxl
from main.models import SlangModel

path = "/home/yorkin1998/Desktop/Projects/Sleng_Dissertatsiya/slanglist.xlsx"


# print("OLDIN:",umumiy)
# print("Qoshilgani:",counter)
wb_obj = openpyxl.load_workbook(path,data_only=True)

sheet_obj = wb_obj.active

for i in range(2,1000):
    if sheet_obj.cell(row=i, column=1).value:
        SlangModel.objects.create(
            slang = sheet_obj.cell(row=i, column=1).value,
            definition = sheet_obj.cell(row=i, column=2).value,
        )
    else:
        continue
