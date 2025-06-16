import os
import django

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
os.sys.path.append(BASE_DIR)
os.environ['DJANGO_SETTINGS_MODULE'] = 'core.settings'
os.environ.setdefault("DJANGO_SETTINGS_MODULE", "core.settings")
django.setup()

import openpyxl
from main.models import SlangSentence

path = "/home/yorkin1998/Desktop/Projects/Sleng_Dissertatsiya/6.xlsx"


# print("OLDIN:",umumiy)
# print("Qoshilgani:",counter)
wb_obj = openpyxl.load_workbook(path,data_only=True)

sheet_obj = wb_obj.active
print("Boshlandi")
for i in range(1,3000):
    if (sheet_obj.cell(row=i, column=1).value) and (sheet_obj.cell(row=i, column=3).value == 1):
        SlangSentence.objects.create(
            slang = sheet_obj.cell(row=i, column=1).value,
            slang_word = sheet_obj.cell(row=i, column=2).value,
        )
    else:
        continue
print("Tugadi")